import os
import secrets
import logging
from datetime import date, datetime
from io import BytesIO

import pandas as pd
from flask import (
    Flask, render_template, request, redirect, url_for,
    flash, send_from_directory, send_file
)
from flask_login import login_required
from openpyxl import Workbook
from openpyxl import load_workbook
from markupsafe import escape
import html
from sqlalchemy import select, or_, func
from sqlalchemy.orm import selectinload
from werkzeug.security import generate_password_hash
from werkzeug.utils import secure_filename

from auth import auth_bp, login_manager
from db import Base, engine, SessionLocal
from models import RFQ, Offerta, Document, User
from utils import export_offerte_excel, render_pdf_from_html

logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
)


# ----------------------------
# Helpers date parsing
# ----------------------------
def parse_date(val: str | None) -> date:
    """Converte 'YYYY-MM-DD' in datetime.date; se vuoto/errato -> oggi."""
    if not val:
        return date.today()
    try:
        return datetime.strptime(val, "%Y-%m-%d").date()
    except ValueError:
        return date.today()


def parse_date_nullable(val: str | None):
    """Come parse_date ma ritorna None se vuoto (utile nei filtri)."""
    if not val:
        return None
    try:
        return datetime.strptime(val, "%Y-%m-%d").date()
    except ValueError:
        return None


# ----------------------------
# Flask App setup
# ----------------------------
app = Flask(__name__)

_secret_key = os.getenv("SECRET_KEY", "dev-key")
if _secret_key == "dev-key":
    logging.warning(
        "SECRET_KEY non impostata — uso il valore di default non sicuro. "
        "Imposta SECRET_KEY nel file .env prima di andare in produzione."
    )

app.config.from_mapping(
    SECRET_KEY=_secret_key,
    # Upload fuori da static/ per evitare accesso diretto senza autenticazione
    UPLOAD_FOLDER=os.getenv(
        "UPLOAD_FOLDER",
        os.path.join(os.path.dirname(os.path.abspath(__file__)), "uploads"),
    ),
    DELETE_AUTH_CODE=os.getenv("DELETE_AUTH_CODE", "DEL2025"),
)

login_manager.init_app(app)
app.register_blueprint(auth_bp, url_prefix="")

# ----------------------------
# DB Init + Admin
# ----------------------------
Base.metadata.create_all(engine)
with SessionLocal() as db:
    if not db.execute(select(User).where(User.username == "admin")).scalar_one_or_none():
        admin_pwd = os.getenv("ADMIN_PASSWORD")
        if not admin_pwd:
            admin_pwd = secrets.token_urlsafe(16)
            print("\n" + "=" * 60)
            print("PRIMO AVVIO — credenziali admin generate automaticamente:")
            print(f"  Username : admin")
            print(f"  Password : {admin_pwd}")
            print("Cambia la password subito da Impostazioni > Cambia password!")
            print("=" * 60 + "\n")
        db.add(User(username="admin", password_hash=generate_password_hash(admin_pwd)))
        db.commit()


# ----------------------------
# Pipeline helper
# ----------------------------
def _pipeline_per_anno(stato_rfq: str, db) -> dict:
    """Ritorna {anno_sop: K€} per le RFQ con lo stato dato, usando l'offerta attiva più recente."""
    rfqs = db.execute(
        select(RFQ)
        .where(RFQ.stato == stato_rfq)
        .where(RFQ.anno_sop.isnot(None))
    ).scalars().all()

    pipeline: dict[int, float] = {}
    for rfq in rfqs:
        latest = db.execute(
            select(Offerta)
            .where(Offerta.rfq_id == rfq.id)
            .where(Offerta.stato == "attiva")
            .order_by(Offerta.id_offerta_rev.desc())
            .limit(1)
        ).scalar_one_or_none()
        if latest:
            fatt_total = sum(v or 0 for v in [
                latest.fatt_sop, latest.fatt_sop1, latest.fatt_sop2,
                latest.fatt_sop3, latest.fatt_sop4,
            ])
            pipeline[rfq.anno_sop] = pipeline.get(rfq.anno_sop, 0) + fatt_total / 1000
    return dict(sorted(pipeline.items()))


# ----------------------------
# Utility per rimozione file/allegati
# ----------------------------
def _safe_remove_file(path: str):
    """Rimuove un file solo se si trova dentro UPLOAD_FOLDER."""
    try:
        if not path:
            return
        abs_root = os.path.abspath(app.config["UPLOAD_FOLDER"])
        abs_path = os.path.abspath(path)
        if abs_path.startswith(abs_root) and os.path.exists(abs_path):
            os.remove(abs_path)
    except Exception:
        # non blocco l'operazione se la cancellazione del file fallisce
        pass


def _delete_offerta_and_docs(db, offerta: Offerta):
    """Cancella un'offerta con i suoi documenti fisici + record."""
    # elimina documenti legati all'offerta
    docs = db.execute(select(Document).where(Document.offerta_id == offerta.id)).scalars().all()
    for d in docs:
        _safe_remove_file(d.path)
        db.delete(d)
    # elimina offerta
    db.delete(offerta)


def _delete_rfq_offerte_and_docs(db, rfq: RFQ):
    """Cancella una RFQ, le offerte collegate e tutti i relativi documenti (RFQ e Offerte)."""
    # Documenti RFQ
    rfq_docs = db.execute(select(Document).where(Document.rfq_id == rfq.id)).scalars().all()
    for d in rfq_docs:
        _safe_remove_file(d.path)
        db.delete(d)

    # Offerte + relativi documenti
    offerte = db.execute(select(Offerta).where(Offerta.rfq_id == rfq.id)).scalars().all()
    for o in offerte:
        _delete_offerta_and_docs(db, o)

    # infine la RFQ
    db.delete(rfq)


# ----------------------------
# Dashboard
# ----------------------------
@app.route("/")
@login_required
def dashboard():
    anno_filter = request.args.get("anno", "").strip()
    anno_int: int | None = None
    date_start: date | None = None
    date_end: date | None = None
    if anno_filter:
        try:
            anno_int = int(anno_filter)
            date_start = date(anno_int, 1, 1)
            date_end   = date(anno_int, 12, 31)
        except ValueError:
            anno_filter = ""

    with SessionLocal() as db:

        def _rfq_count(*extra):
            stmt = select(func.count()).select_from(RFQ)
            if date_start:
                stmt = stmt.where(RFQ.data_ricezione >= date_start).where(RFQ.data_ricezione <= date_end)
            for w in extra:
                stmt = stmt.where(w)
            return db.scalar(stmt) or 0

        total_rfq_ricevute = _rfq_count()
        total_rfq_attive   = _rfq_count(RFQ.stato == "attiva")

        off_stmt = select(func.count()).select_from(Offerta)
        if date_start:
            off_stmt = off_stmt.where(Offerta.data_offerta >= date_start).where(Offerta.data_offerta <= date_end)
        total_offerte = db.scalar(off_stmt) or 0

        stati = ("attiva", "inattiva", "vinta", "persa", "non_gestita", "obsoleta")
        rfq_status_counts = {s: _rfq_count(RFQ.stato == s) for s in stati}

        # RFQ recenti (filtrate per anno se selezionato, escluse obsolete)
        rfq_recent_stmt = select(RFQ).where(RFQ.stato != "obsoleta")
        if date_start:
            rfq_recent_stmt = rfq_recent_stmt.where(
                RFQ.data_ricezione >= date_start).where(RFQ.data_ricezione <= date_end)
        recent_rfq = db.execute(rfq_recent_stmt.order_by(RFQ.id.desc()).limit(5)).scalars().all()

        # Offerte recenti (filtrate per anno se selezionato)
        off_recent_stmt = select(Offerta).options(selectinload(Offerta.rfq))
        if date_start:
            off_recent_stmt = off_recent_stmt.where(
                Offerta.data_offerta >= date_start).where(Offerta.data_offerta <= date_end)
        recent_off = db.execute(off_recent_stmt.order_by(Offerta.id.desc()).limit(5)).scalars().all()

        # Pipeline: sempre globale (basata su anno_sop, non sulla data ricezione)
        pipeline_attive = _pipeline_per_anno("attiva", db)
        pipeline_vinte  = _pipeline_per_anno("vinta",  db)

    return render_template(
        "dashboard.html",
        anno=anno_filter,
        total_rfq_ricevute=total_rfq_ricevute,
        total_rfq_attive=total_rfq_attive,
        total_offerte=total_offerte,
        rfq_status_counts=rfq_status_counts,
        recent_rfq=recent_rfq,
        recent_off=recent_off,
        pipeline_attive=pipeline_attive,
        pipeline_vinte=pipeline_vinte,
    )


# ----------------------------
# RFQ CRUD
# ----------------------------
@app.route("/rfq")
@login_required
def rfq_list():
    q = request.args.get("q", "").strip()
    anno_filter = request.args.get("anno", "").strip()
    mostra_obsolete = request.args.get("mostra_obsolete", "0") == "1"

    with SessionLocal() as db:
        stmt = select(RFQ)

        if q:
            like = f"%{q}%"
            stmt = stmt.where(or_(RFQ.nome_cliente.ilike(like), RFQ.nome_progetto.ilike(like)))

        if anno_filter:
            try:
                anno_int = int(anno_filter)
                start = date(anno_int, 1, 1)
                end = date(anno_int, 12, 31)
                stmt = stmt.where(RFQ.data_ricezione >= start).where(RFQ.data_ricezione <= end)
            except ValueError:
                anno_filter = ""

        if not mostra_obsolete:
            stmt = stmt.where(RFQ.stato != "obsoleta")

        rfqs = db.execute(stmt.order_by(RFQ.nome_cliente, RFQ.nome_progetto)).scalars().all()

        # Statistiche anno
        anno_stats = None
        if anno_filter:
            anno_int = int(anno_filter)
            start_d = date(anno_int, 1, 1)
            end_d = date(anno_int, 12, 31)
            rfqs_anno = db.execute(
                select(RFQ).where(RFQ.data_ricezione >= start_d).where(RFQ.data_ricezione <= end_d)
            ).scalars().all()
            by_stato: dict[str, int] = {}
            for r in rfqs_anno:
                by_stato[r.stato] = by_stato.get(r.stato, 0) + 1
            offerte_anno = db.scalar(
                select(func.count()).select_from(Offerta)
                .where(Offerta.data_offerta >= start_d)
                .where(Offerta.data_offerta <= end_d)
            ) or 0
            anno_stats = {
                "anno": anno_int,
                "total_rfq": len(rfqs_anno),
                "total_offerte": offerte_anno,
                "by_stato": by_stato,
            }

    return render_template(
        "rfq_list.html",
        rfqs=rfqs, q=q,
        anno=anno_filter,
        mostra_obsolete=mostra_obsolete,
        anno_stats=anno_stats,
    )


@app.route("/rfq/new", methods=["GET", "POST"])
@login_required
def rfq_new():
    if request.method == "POST":
        with SessionLocal() as db:
            stato = request.form.get("stato", "attiva")
            anno_sop_s = request.form.get("anno_sop", "").strip()
            r = RFQ(
                nome_cliente=request.form["nome_cliente"].strip(),
                nome_progetto=request.form["nome_progetto"].strip(),
                data_ricezione=parse_date(request.form.get("data_ricezione")),
                due_date_quotazione=parse_date(request.form.get("due_date_quotazione")),
                target_price=float(request.form["target_price"]) if request.form.get("target_price") else None,
                descrizione=request.form.get("descrizione"),
                stato=stato,
                anno_sop=int(anno_sop_s) if anno_sop_s.isdigit() else None,
                motivo_non_gestione=request.form.get("motivo_non_gestione") if stato == "non_gestita" else None,
            )
            db.add(r)
            try:
                db.commit()
                flash("RFQ creata correttamente ✅", "success")
                return redirect(url_for("rfq_list"))
            except Exception as e:
                db.rollback()
                flash(f"Errore durante il salvataggio: {e}", "danger")
    return render_template("rfq_form.html", item=None)


@app.route("/rfq/<int:rfq_id>/edit", methods=["GET", "POST"])
@login_required
def rfq_edit(rfq_id):
    with SessionLocal() as db:
        r = db.get(RFQ, rfq_id)
        if not r:
            flash("RFQ non trovata ❌", "warning")
            return redirect(url_for("rfq_list"))
        if request.method == "POST":
            stato = request.form.get("stato", "attiva")
            anno_sop_s = request.form.get("anno_sop", "").strip()
            r.nome_cliente = request.form["nome_cliente"].strip()
            r.nome_progetto = request.form["nome_progetto"].strip()
            r.data_ricezione = parse_date(request.form.get("data_ricezione"))
            r.due_date_quotazione = parse_date(request.form.get("due_date_quotazione"))
            r.target_price = float(request.form["target_price"]) if request.form.get("target_price") else None
            r.descrizione = request.form.get("descrizione")
            r.stato = stato
            r.anno_sop = int(anno_sop_s) if anno_sop_s.isdigit() else None
            r.motivo_non_gestione = request.form.get("motivo_non_gestione") if stato == "non_gestita" else None
            try:
                db.commit()
                flash("RFQ aggiornata correttamente ✅", "success")
                return redirect(url_for("rfq_list"))
            except Exception as e:
                db.rollback()
                flash(f"Errore durante l'aggiornamento: {e}", "danger")
    return render_template("rfq_form.html", item=r)


@app.route("/rfq/<int:rfq_id>")
@login_required
def rfq_detail(rfq_id):
    with SessionLocal() as db:
        r = db.execute(
            select(RFQ)
            .options(
                selectinload(RFQ.documents),
                selectinload(RFQ.offerte).selectinload(Offerta.documents)
            )
            .where(RFQ.id == rfq_id)
        ).scalar_one_or_none()

        if not r:
            flash("RFQ non trovata", "warning")
            return redirect(url_for("rfq_list"))

        offerte = sorted(r.offerte, key=lambda o: o.id_offerta_rev, reverse=True)

    return render_template("rfq_detail.html", rfq=r, offerte=offerte, docs=r.documents)


@app.route("/rfq/<int:rfq_id>/set_state", methods=["POST"])
@login_required
def rfq_set_state(rfq_id):
    """Aggiorna rapidamente lo stato della RFQ (attiva, inattiva, vinta, persa)."""
    new_state = request.form.get("stato")

    if new_state not in ("attiva", "inattiva", "vinta", "persa", "non_gestita", "obsoleta"):
        flash("⚠️ Stato non valido.", "warning")
        return redirect(url_for("rfq_detail", rfq_id=rfq_id))

    with SessionLocal() as db:
        r = db.get(RFQ, rfq_id)
        if not r:
            flash("RFQ non trovata", "warning")
            return redirect(url_for("rfq_list"))
        r.stato = new_state
        db.commit()

    flash(f"✅ Stato RFQ aggiornato a: {new_state}", "success")
    return redirect(url_for("rfq_detail", rfq_id=rfq_id))

@app.route("/rfq/<int:rfq_id>/delete", methods=["POST"])
@login_required
def rfq_delete(rfq_id):
    code = request.form.get("auth_code", "").strip()
    if code != app.config["DELETE_AUTH_CODE"]:
        flash("Codice autorizzativo non valido ❌", "danger")
        return redirect(request.referrer or url_for("rfq_list"))

    with SessionLocal() as db:
        r = db.get(RFQ, rfq_id)
        if not r:
            flash("RFQ non trovata", "warning")
            return redirect(url_for("rfq_list"))

        _delete_rfq_offerte_and_docs(db, r)
        db.commit()

    flash("RFQ, offerte e allegati eliminati ✅", "success")
    return redirect(url_for("rfq_list"))


@app.route("/rfq/<int:rfq_id>/clone_rev", methods=["POST"])
@login_required
def rfq_clone_rev(rfq_id):
    """Crea una nuova revisione della RFQ (dati clonati, nessuna offerta) e mette l'originale in OBSOLETA."""
    with SessionLocal() as db:
        r = db.get(RFQ, rfq_id)
        if not r:
            flash("RFQ non trovata", "warning")
            return redirect(url_for("rfq_list"))
        if r.stato == "obsoleta":
            flash("Non puoi creare una revisione di una RFQ già obsoleta.", "warning")
            return redirect(url_for("rfq_detail", rfq_id=rfq_id))

        new_rev = r.numero_revisione + 1
        new_rfq = RFQ(
            nome_cliente=r.nome_cliente,
            nome_progetto=r.nome_progetto,
            data_ricezione=date.today(),
            due_date_quotazione=r.due_date_quotazione,
            target_price=r.target_price,
            descrizione=r.descrizione,
            stato="attiva",
            anno_sop=r.anno_sop,
            motivo_non_gestione=None,
            numero_revisione=new_rev,
            rfq_padre_id=r.id,
        )
        r.stato = "obsoleta"
        db.add(new_rfq)
        try:
            db.commit()
            db.refresh(new_rfq)
            flash(f"✅ Creata revisione Rev.{new_rev} — la precedente è stata impostata come OBSOLETA", "success")
            return redirect(url_for("rfq_detail", rfq_id=new_rfq.id))
        except Exception as e:
            db.rollback()
            flash(f"Errore durante la clonazione: {e}", "danger")
            return redirect(url_for("rfq_detail", rfq_id=rfq_id))


# ----------------------------
# Offerte CRUD
# ----------------------------
def _offerta_from_form(o: Offerta, form):
    o.codice_pm = form.get("codice_pm", "").strip()
    o.id_offerta_rev = int(form.get("id_offerta_rev", "1") or 1)
    o.versione = form.get("versione") or None
    o.stato = form.get("stato", "attiva")
    o.data_offerta = parse_date(form.get("data_offerta"))

    for fld in [
        "prezzo_sop", "prezzo_sop1", "prezzo_sop2", "prezzo_sop3", "prezzo_sop4",
        "gm_sop", "gm_sop1", "gm_sop2", "gm_sop3", "gm_sop4",
    ]:
        v = form.get(fld)
        setattr(o, fld, float(v) if v not in (None, "") else None)

    for fld in ["vol_sop", "vol_sop1", "vol_sop2", "vol_sop3", "vol_sop4"]:
        v = form.get(fld)
        setattr(o, fld, int(v) if v not in (None, "") else None)
    return o


@app.route("/rfq/<int:rfq_id>/offerta/new", methods=["GET", "POST"])
@login_required
def offerta_new(rfq_id):
    with SessionLocal() as db:
        r = db.get(RFQ, rfq_id)
        if not r:
            flash("RFQ non trovata", "warning")
            return redirect(url_for("rfq_list"))

        if request.method == "POST":
            o = Offerta(rfq_id=rfq_id, codice_pm="tmp", id_offerta_rev=1, stato="attiva")
            o = _offerta_from_form(o, request.form)
            db.add(o)
            try:
                db.commit()
                db.refresh(o)  # assicura che o.id sia popolato
                flash("Offerta creata correttamente ✅", "success")
                return redirect(url_for("rfq_detail", rfq_id=rfq_id))
            except Exception as e:
                db.rollback()
                flash(f"Errore durante il salvataggio: {e}", "danger")

        max_rev = db.scalar(
            select(func.max(Offerta.id_offerta_rev)).where(Offerta.rfq_id == rfq_id)
        ) or 0
    return render_template("offerta_form.html", rfq_id=rfq_id, item=None, suggested_rev=max_rev + 1)


@app.route("/offerta/<int:off_id>/edit", methods=["GET", "POST"])
@login_required
def offerta_edit(off_id):
    with SessionLocal() as db:
        o = db.get(Offerta, off_id)
        if not o:
            flash("Offerta non trovata ❌", "warning")
            return redirect(url_for("rfq_list"))

        if request.method == "POST":
            _offerta_from_form(o, request.form)
            try:
                db.commit()
                flash("Offerta aggiornata correttamente ✅", "success")
                return redirect(url_for("rfq_detail", rfq_id=o.rfq_id))
            except Exception as e:
                db.rollback()
                flash(f"Errore durante l'aggiornamento: {e}", "danger")

    return render_template("offerta_form.html", rfq_id=o.rfq_id, item=o, suggested_rev=o.id_offerta_rev)


@app.route("/offerta/<int:off_id>/delete", methods=["POST"])
@login_required
def offerta_delete(off_id):
    code = request.form.get("auth_code", "").strip()
    if code != app.config["DELETE_AUTH_CODE"]:
        flash("Codice autorizzativo non valido ❌", "danger")
        return redirect(request.referrer or url_for("dashboard"))

    with SessionLocal() as db:
        o = db.get(Offerta, off_id)
        if not o:
            flash("Offerta non trovata", "warning")
            return redirect(url_for("rfq_list"))

        rfq_id = o.rfq_id
        _delete_offerta_and_docs(db, o)
        db.commit()

    flash("Offerta e relativi allegati eliminati ✅", "success")
    return redirect(url_for("rfq_detail", rfq_id=rfq_id))


# ----------------------------
# Import / Export Excel (Offerta)
# ----------------------------
@app.route("/download/offerta_template.xlsx")
@login_required
def download_offerta_template():
    wb = Workbook()
    ws = wb.active
    ws.title = "Offerta Template"
    ws.append([
        "Codice_PM", "Revisione", "Versione", "Stato", "Data_Offerta",
        "Prezzo_SOP", "Volumi_SOP", "GM_SOP",
        "Prezzo_SOP1", "Volumi_SOP1", "GM_SOP1",
        "Prezzo_SOP2", "Volumi_SOP2", "GM_SOP2",
        "Prezzo_SOP3", "Volumi_SOP3", "GM_SOP3",
        "Prezzo_SOP4", "Volumi_SOP4", "GM_SOP4",
    ])
    out = BytesIO(); wb.save(out); out.seek(0)
    return send_file(out, as_attachment=True, download_name="offerta_template.xlsx",
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


@app.route("/rfq/<int:rfq_id>/import_excel", methods=["POST"])
@login_required
def import_offerta_excel(rfq_id):
    file = request.files.get("file")
    if not file:
        flash("Nessun file selezionato per l'import.", "warning")
        return redirect(request.referrer)

    try:
        df = pd.read_excel(file)
        row = df.iloc[0].to_dict()
    except Exception as e:
        flash(f"Errore durante la lettura del file Excel: {e}", "danger")
        return redirect(request.referrer)

    with SessionLocal() as db:
        # Calcola revisione successiva per questa RFQ
        max_rev = db.scalar(
            select(func.max(Offerta.id_offerta_rev)).where(Offerta.rfq_id == rfq_id)
        ) or 0

        o = Offerta(
            rfq_id=rfq_id,
            codice_pm=row.get("Codice_PM", "N/A"),
            id_offerta_rev=max_rev + 1,
            versione=row.get("Versione") or f"R{max_rev + 1}",
            stato=row.get("Stato", "attiva"),
            data_offerta=parse_date(str(row.get("Data_Offerta", date.today()))),
        )
        for i in ["", "1", "2", "3", "4"]:
            setattr(o, f"prezzo_sop{i}", float(row.get(f"Prezzo_SOP{i}", 0)))
            setattr(o, f"vol_sop{i}", int(row.get(f"Volumi_SOP{i}", 0)))
            setattr(o, f"gm_sop{i}", float(row.get(f"GM_SOP{i}", 0)))

        db.add(o)
        db.commit()

    flash(f"Offerta importata correttamente come revisione R{max_rev + 1} ✅", "success")
    return redirect(url_for("rfq_detail", rfq_id=rfq_id))


# ----------------------------
# Offerte list + Export
# ----------------------------
@app.route("/offerte")
@login_required
def offerte_list():
    cliente = request.args.get("cliente", "").strip()
    progetto = request.args.get("progetto", "").strip()
    dal_str = request.args.get("dal", "")
    al_str = request.args.get("al", "")
    dal = parse_date_nullable(dal_str)
    al = parse_date_nullable(al_str)

    query = select(Offerta).join(RFQ)
    if cliente:
        query = query.where(RFQ.nome_cliente.ilike(f"%{cliente}%"))
    if progetto:
        query = query.where(RFQ.nome_progetto.ilike(f"%{progetto}%"))
    if dal:
        query = query.where(Offerta.data_offerta >= dal)
    if al:
        query = query.where(Offerta.data_offerta <= al)

    with SessionLocal() as db:
        res = db.execute(
            query.options(selectinload(Offerta.rfq)).order_by(Offerta.data_offerta.desc())
        ).scalars().all()

    return render_template(
        "offerte_list.html",
        items=res,
        f=dict(cliente=cliente, progetto=progetto, dal=dal_str, al=al_str),
    )


@app.route("/export/offerte.xlsx")
@login_required
def export_offerte_xlsx():
    with SessionLocal() as db:
        rows = db.execute(
            select(Offerta)
            .options(selectinload(Offerta.rfq))
            .order_by(Offerta.data_offerta.desc())
        ).scalars().all()
    return export_offerte_excel(rows)


@app.route("/export/offerte.pdf")
@login_required
def export_offerte_pdf():
    with SessionLocal() as db:
        rows = db.execute(
            select(Offerta)
            .options(selectinload(Offerta.rfq))
            .order_by(Offerta.data_offerta.desc())
        ).scalars().all()
    html = render_template("offerte_pdf.html", items=rows)
    return render_pdf_from_html(html, filename="offerte.pdf")


# ----------------------------
# Upload, Files, Clone
# ----------------------------
@app.route("/upload", methods=["POST"])
@login_required
def upload():
    file = request.files.get("file")
    rfq_id = request.form.get("rfq_id")
    offerta_id = request.form.get("offerta_id")
    if not file or file.filename == "":
        flash("Nessun file selezionato", "warning")
        return redirect(request.referrer or url_for("dashboard"))

    safe_name = secure_filename(file.filename)
    if not safe_name:
        flash("Nome file non valido", "warning")
        return redirect(request.referrer or url_for("dashboard"))
    os.makedirs(app.config["UPLOAD_FOLDER"], exist_ok=True)
    save_path = os.path.join(app.config["UPLOAD_FOLDER"], safe_name)
    file.save(save_path)

    with SessionLocal() as db:
        d = Document(
            rfq_id=int(rfq_id) if rfq_id else None,
            offerta_id=int(offerta_id) if offerta_id else None,
            filename=safe_name,
            path=save_path,
        )
        db.add(d)
        db.commit()

    flash("File caricato ✅", "success")
    return redirect(request.referrer or url_for("dashboard"))


@app.route("/files/<path:filename>")
@login_required
def files(filename):
    return send_from_directory(app.config["UPLOAD_FOLDER"], filename, as_attachment=True)

@app.route("/view_excel/<path:filename>")
@login_required
def view_excel(filename):
    import os
    filepath = os.path.join(app.config["UPLOAD_FOLDER"], filename)
    if not os.path.exists(filepath):
        flash("File non trovato.", "danger")
        return redirect(request.referrer or url_for("dashboard"))

    return render_template("view_excel.html", filename=filename)

# ----------------------------
# Visualizza Excel "Payback" server-side
# ----------------------------
from openpyxl import load_workbook
from markupsafe import Markup

@app.route("/view_excel_server/<path:filename>")
@login_required
def view_excel_server(filename):
    """Mostra la scheda Payback_Values (se c'è) altrimenti la prima Payback."""
    import re
    from openpyxl import load_workbook
    from markupsafe import Markup
    from markupsafe import escape
    import html

    file_path = os.path.join(app.config["UPLOAD_FOLDER"], filename)
    if not os.path.exists(file_path):
        flash("File non trovato ❌", "danger")
        return redirect(request.referrer or url_for("dashboard"))

    # normalizza i nomi foglio (leva spazi, underscore, trattini; tutto minuscolo)
    def norm(s: str) -> str:
        return re.sub(r"[^a-z0-9]", "", s.lower())

    try:
        wb = load_workbook(file_path, data_only=True)
        names = wb.sheetnames
        norm_map = {name: norm(name) for name in names}

        # 1) prova con payback_values
        preferred = [n for n in names if "paybackvalues" in norm_map[n]]
        # 2) altrimenti qualsiasi payback
        fallback = [n for n in names if "payback" in norm_map[n]]

        if preferred:
            sheet_name = preferred[0]
        elif fallback:
            sheet_name = fallback[0]
        else:
            wb.close()
            flash("❌ Nessuna scheda 'Payback' trovata nel file.", "warning")
            return redirect(request.referrer or url_for("dashboard"))

        ws = wb[sheet_name]

        # costruisci tabella html semplice
        rows_html = []
        for row in ws.iter_rows(values_only=True):
            tds = "".join(
                f"<td style='border:1px solid #ddd; padding:4px 6px; font-size:13px'>{'' if v is None else v}</td>"
                for v in row
            )
            rows_html.append(f"<tr>{tds}</tr>")
        table_html = "<table class='thin' style='border-collapse:collapse; width:100%'>" + "".join(rows_html) + "</table>"

        wb.close()
        return render_template(
            "view_excel_server.html",
            filename=filename,
            sheet_name=sheet_name,            # <- visibile nella pagina
            table_html=Markup(table_html)
        )
    except Exception as e:
        flash(f"Errore nella lettura del file Excel: {e}", "danger")
        return redirect(request.referrer or url_for("dashboard"))

@app.route("/view_excel_styled/<path:filename>")
@login_required
def view_excel_styled(filename):
    """
    Visualizza in HTML la prima scheda che CONTIENE 'payback' nel nome
    ma ignora (1) i tab che si chiamano ESATTAMENTE 'Payback' e (2) i tab con 'Example' nel nome.
    Rendering lato server con openpyxl (colori, font, allineamenti basilari, number format).
    """
    # Sicurezza: consenti solo file dentro UPLOAD_FOLDER
    upload_root = os.path.abspath(app.config["UPLOAD_FOLDER"])
    safe_name = os.path.normpath(filename)
    full_path = os.path.abspath(os.path.join(upload_root, safe_name))
    if not full_path.startswith(upload_root) or not os.path.exists(full_path):
        flash("File non trovato o percorso non valido.", "danger")
        return redirect(request.referrer or url_for("dashboard"))

    try:
        wb = load_workbook(full_path, data_only=True)  # data_only=True = valori calcolati (non formule)
    except Exception as e:
        flash(f"Errore apertura Excel: {e}", "danger")
        return redirect(request.referrer or url_for("dashboard"))

    # Selezione foglio: contiene 'payback' ma NON (nome == 'payback') e NON contiene 'example'
    chosen_ws = None
    for name in wb.sheetnames:
        lower = name.lower()
        if ("payback" in lower) and (lower != "payback") and ("example" not in lower):
            chosen_ws = wb[name]
            break

    if chosen_ws is None:
        flash("Nessuna scheda valida trovata (contiene 'Payback' ma non 'Example' e non esattamente 'Payback').", "warning")
        return redirect(request.referrer or url_for("dashboard"))

    # Helper: formattazione numerica *semplice* basata su number_format di Excel
    def fmt_value(cell):
        v = cell.value
        if v is None:
            return ""
        # Se è già stringa -> escapa e restituisci
        if isinstance(v, str):
            return html.escape(v)

        nf = (cell.number_format or "").lower()

        try:
            if isinstance(v, (int, float)):
                # percentuali
                if "%" in nf:
                    return f"{(v * 100):.2f}%"
                # valuta (eur, usd ecc)
                if "€" in nf or "[$" in nf or "€" in nf:
                    # migliaia + 2 decimali
                    return f"€ {v:,.2f}".replace(",", " ").replace(".", ",").replace(" ", ".")
                # migliaia / decimali
                if "0," in nf or "0.00" in nf:
                    return f"{v:,.2f}".replace(",", " ").replace(".", ",").replace(" ", ".")
                # numero intero
                if "0" in nf and ".00" not in nf and "," not in nf:
                    return f"{int(round(v))}"
                # default: numero con max 2 decimali
                return f"{v:,.2f}".replace(",", " ").replace(".", ",").replace(" ", ".")
        except Exception:
            pass

        return html.escape(str(v))

    # Cattura range effettivo con dati
    min_row = chosen_ws.min_row
    max_row = chosen_ws.max_row
    min_col = chosen_ws.min_column
    max_col = chosen_ws.max_column

    # Costruzione HTML tabella con stile minimo + colori
    rows_html = []
    for r in range(min_row, max_row + 1):
        tds = []
        for c in range(min_col, max_col + 1):
            cell = chosen_ws.cell(row=r, column=c)
            text = fmt_value(cell)

            # stile base
            styles = []

            # background (fill)
            fill = cell.fill
            bg = None
            if fill and getattr(fill, "fgColor", None):
                try:
                    rgb = getattr(fill.fgColor, "rgb", None)
                    if isinstance(rgb, str):
                        if len(rgb) == 8:
                            bg = f"#{rgb[2:]}"
                        elif len(rgb) == 6:
                            bg = f"#{rgb}"
                    elif hasattr(fill.fgColor, "theme") or hasattr(fill.fgColor, "tint"):
                        # Colore tema o gradiente — fallback neutro
                        bg = "#f8f9fa"
                except Exception:
                    pass
            if bg and bg.lower() != "#000000":
                styles.append(f"background:{bg};")

            # font
            font = cell.font
            if font:
                if font.bold:
                    styles.append("font-weight:bold;")
                if font.italic:
                    styles.append("font-style:italic;")

                # Colore testo — gestione robusta anche per oggetti RGB/Color
                if font.color:
                    try:
                        frgb = getattr(font.color, "rgb", None)
                        if isinstance(frgb, str):
                            if len(frgb) == 8:
                                styles.append(f"color:#{frgb[2:]};")
                            elif len(frgb) == 6:
                                styles.append(f"color:#{frgb};")
                        elif hasattr(font.color, "theme") or hasattr(font.color, "tint"):
                            # Fallback colore tema
                            styles.append("color:#333;")
                    except Exception:
                        pass

            # align
            align = cell.alignment
            if align and align.horizontal:
                styles.append(f"text-align:{align.horizontal};")
            else:
                # default: numeri a destra, altro a sinistra
                styles.append("text-align:right;" if isinstance(cell.value, (int, float)) else "text-align:left;")

            # bordi molto leggeri
            styles.append("border:1px solid #ddd; padding:4px 6px;")

            style_attr = f' style="{"".join(styles)}"' if styles else ""
            tds.append(f"<td{style_attr}>{text}</td>")
        rows_html.append(f"<tr>{''.join(tds)}</tr>")

    table_html = (
        '<table style="border-collapse:collapse; width:100%; font-size:13px; background:#fff;">'
        f"{''.join(rows_html)}"
        "</table>"
    )

    return render_template(
        "view_excel_styled.html",
        filename=filename,
        sheet_name=chosen_ws.title,
        table_html=table_html,
    )

@app.route("/offerta/<int:off_id>/clone_rev", methods=["POST"])
@login_required
def offerta_clone_rev(off_id):
    with SessionLocal() as db:
        o = db.get(Offerta, off_id)
        if not o:
            flash("Offerta non trovata", "warning")
            return redirect(url_for("rfq_list"))

        max_rev = db.scalar(
            select(func.max(Offerta.id_offerta_rev)).where(Offerta.rfq_id == o.rfq_id)
        ) or 0

        new_o = Offerta(
            rfq_id=o.rfq_id,
            codice_pm=o.codice_pm,
            id_offerta_rev=max_rev + 1,
            versione=f"R{max_rev + 1}",
            stato="attiva",
            data_offerta=date.today(),
            prezzo_sop=o.prezzo_sop,
            prezzo_sop1=o.prezzo_sop1,
            prezzo_sop2=o.prezzo_sop2,
            prezzo_sop3=o.prezzo_sop3,
            prezzo_sop4=o.prezzo_sop4,
            gm_sop=o.gm_sop,
            gm_sop1=o.gm_sop1,
            gm_sop2=o.gm_sop2,
            gm_sop3=o.gm_sop3,
            gm_sop4=o.gm_sop4,
            vol_sop=o.vol_sop,
            vol_sop1=o.vol_sop1,
            vol_sop2=o.vol_sop2,
            vol_sop3=o.vol_sop3,
            vol_sop4=o.vol_sop4,
        )
        db.add(new_o)
        db.commit()

    flash(f"Creata revisione R{max_rev + 1} ✅", "success")
    return redirect(url_for("rfq_detail", rfq_id=o.rfq_id))


# ----------------------------
# Context processors
# ----------------------------
@app.context_processor
def inject_now():
    return {"now": datetime.now}


@app.context_processor
def inject_date():
    return {"date": date}


# ----------------------------
# Run (solo per sviluppo locale — in produzione usa run.py con Waitress)
# ----------------------------
if __name__ == "__main__":
    app.run(debug=False, host="127.0.0.1", port=8080)